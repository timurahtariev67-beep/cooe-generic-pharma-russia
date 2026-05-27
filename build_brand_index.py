#!/usr/bin/env python3
"""
Build BrandIndex sheet and add Brand_norm / BrandIndex_lag1 to основа_норм.

BrandIndex = (YearsScore + CoverageScore + PortfolioScore + ProducerTypeScore) / 4

YearsScore    – years since first observation in dataset (proxy for registration date)
CoverageScore – brand pack-share in year t, mapped to [0-5] using rescaled thresholds
PortfolioScore – unique canonical brand names per manufacturer per year → [0-5]
ProducerTypeScore – rule-based classification by country + company name
"""

import openpyxl
import math
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

FILE = "/Users/timurahtariev/Desktop/курсовая/курсовая данные новые (2).xlsx"

# ─────────────────────────────────────────────
# 1.  Brand name normalisation
# ─────────────────────────────────────────────

EXPLICIT_BRAND_MAP = {
    # Packaging / applicator accessories → strip to core brand
    "МИРАМИСТИН С АППЛИКАТОРОМ УРОЛОГИЧЕСКИМ (КЛАССИЧЕСКИЙ)":
        "МИРАМИСТИН",
    "МИРАМИСТИН С АППЛИКАТОРОМ УРОЛОГИЧЕСКИМ/С НАСАДКОЙ-РАСПЫЛИТЕЛЕМ (ДОРОЖНАЯ/ТУРИСТИЧЕСКАЯ)":
        "МИРАМИСТИН",
    "МИРАМИСТИН С НАСАДКОЙ ГИНЕКОЛОГИЧЕСКОЙ":
        "МИРАМИСТИН",
    "МИРАМИСТИН С НАСАДКОЙ-РАСПЫЛИТЕЛЕМ":
        "МИРАМИСТИН",
    "МИРАМЕД ЭВАЛАР (НАСАДКА-РАСПЫЛИТЕЛЬ)":
        "МИРАМЕД ЭВАЛАР",
    "МИРАМЕД ЭВАЛАР (С АППЛИКАТОРОМ УРОЛОГИЧЕСКИМ И НАСАДКОЙ ВАГИНАЛЬНОЙ)":
        "МИРАМЕД ЭВАЛАР",
    "ЗИЛЕКСА СЕПТ (ИНТИМ+НАСАДКА/АППЛИКАТОР УРОЛОГИЧЕСКИЙ)":
        "ЗИЛЕКСА СЕПТ",
    "АКВАМАСТЕР ПРОМО+ПОДАРОК (СРЕДСТВО ДЛЯ НОСА МОРСКАЯ СОЛЬ)":
        "АКВАМАСТЕР",
    "БРИЛЛИАНТОВЫЙ ЗЕЛЕНЫЙ С ЛОПАТКОЙ":
        "БРИЛЛИАНТОВЫЙ ЗЕЛЕНЫЙ",
    "ЭНТЕРОЛ (БЛИСТЕР)":
        "ЭНТЕРОЛ",
    "СИАЛОР (ПРОТАРГОЛ) НАБОР ДЛЯ ПРИГОТОВЛЕНИЯ РАСТВОРА (КАПЛИ)":
        "СИАЛОР",
    "СИАЛОР (ПРОТАРГОЛ) НАБОР ДЛЯ ПРИГОТОВЛЕНИЯ РАСТВОРА (СПРЕЙ)":
        "СИАЛОР",
    # Parenthetical clarifications not part of brand identity
    "ВИТАМИН Е (АЛЬФА-ТОКОФЕРОЛА АЦЕТАТ)":
        "ВИТАМИН Е",
    "ЭРГОКАЛЬЦИФЕРОЛ-ЛЕКТ (ВИТАМИН Д2)":
        "ЭРГОКАЛЬЦИФЕРОЛ-ЛЕКТ",
    # Old-style genitive+form+dose names → canonical drug brand name
    "АЛЬФА-ТОКОФЕРОЛА АЦЕТАТА РАСТВОР В МАСЛЕ (ВИТАМИН Е)":
        "АЛЬФА-ТОКОФЕРОЛА АЦЕТАТ",
    "АЛЬФА-ТОКОФЕРОЛА АЦЕТАТА РАСТВОР В МАСЛЕ 50% (ВИТАМИН Е)":
        "АЛЬФА-ТОКОФЕРОЛА АЦЕТАТ",
    "АНАЛЬГИНА РАСТВОР ДЛЯ ИНЪЕКЦИЙ 50%":
        "АНАЛЬГИН",
    "АНАЛЬГИНА ТАБЛЕТКИ 0.5 Г":
        "АНАЛЬГИН",
    "ВАЛИДОЛА ТАБЛЕТКИ 0.06 Г":
        "ВАЛИДОЛ",
    "ВИКАСОЛА ТАБЛЕТКИ 0.015 Г":
        "ВИКАСОЛ",
    "ГЛИЦИРАМА ГРАНУЛЫ ДЛЯ ДЕТЕЙ 0.025 Г":
        "ГЛИЦИРАМА",
    "ДАЛАРГИН ЛИОФИЛИЗИРОВАННЫЙ ДЛЯ ИНЪЕКЦИЙ 0.001 Г":
        "ДАЛАРГИН",
    "ДИМЕФОСФОН ДЛЯ ИНЪЕКЦИЙ 1 Г":
        "ДИМЕФОСФОН",
    "ЙОДА РАСТВОР СПИРТОВОЙ 5%":
        "ЙОД СПИРТОВОЙ",
    "ЙОДА РАСТВОР СПИРТОВОЙ 5% В АМПУЛАХ":
        "ЙОД СПИРТОВОЙ",
    "КАЛИЯ ОРОТАТА ТАБЛЕТКИ 0.5 Г":
        "КАЛИЯ ОРОТАТ",
    "КАЛЬЦИЯ ХЛОРИДА РАСТВОР ДЛЯ ИНЪЕКЦИЙ 10%":
        "КАЛЬЦИЯ ХЛОРИД",
    "КАМФОРНЫЙ СПИРТ 10%":
        "КАМФОРНЫЙ СПИРТ",
    "КЕТОТИФЕНА СИРОП 0.02%":
        "КЕТОТИФЕН",
    "ЛЕВОМИЦЕТИНА РАСТВОР 0.25%":
        "ЛЕВОМИЦЕТИН",
    "МЕТАЦИНА ТАБЛЕТКИ 0.002 Г":
        "МЕТАЦИН",
    "МЕТИЛУРАЦИЛОВАЯ МАЗЬ 10%":
        "МЕТИЛУРАЦИЛ",
    "НЕОВИРА РАСТВОР ДЛЯ ИНЪЕКЦИЙ 12.5%":
        "НЕОВИР",
    "НИЦЕРГОЛИН ДЛЯ ИНЪЕКЦИЙ 0.004 Г":
        "НИЦЕРГОЛИН",
    "НИЦЕРГОЛИНА ТАБЛЕТКИ ПОКРЫТЫЕ ОБОЛОЧКОЙ 0.01 Г":
        "НИЦЕРГОЛИН",
    "НАФТИЗИНА РАСТВОР":
        "НАФТИЗИН",
    "ПАПАВЕРИНА ГИДРОХЛОРИДА ТАБЛЕТКИ 0.04 Г":
        "ПАПАВЕРИНА ГИДРОХЛОРИД",
    "ПАРАЦЕТАМОЛ 325 МГ":
        "ПАРАЦЕТАМОЛ",
    "ПАРАЦЕТАМОЛА СИРОП 2.4%":
        "ПАРАЦЕТАМОЛ",
    "ПЕНТОКСИФИЛЛИНА ТАБЛЕТКИ 0.1 Г (РАСТВОРИМЫЕ В КИШЕЧНИКЕ)":
        "ПЕНТОКСИФИЛЛИН",
    "ПУСТЫРНИКА ЭКСТРАКТА ТАБЛЕТКИ 0.014 Г":
        "ПУСТЫРНИКА ЭКСТРАКТ",
    "РЕФОРТАН ГЭК 10%":
        "РЕФОРТАН ГЭК",
    "РИБОФЛАВИН-МОНОНУКЛЕОТИДА РАСТВОР ДЛЯ ИНЪЕКЦИЙ 1%":
        "РИБОФЛАВИН-МОНОНУКЛЕОТИД",
    "СИБАЗОНА РАСТВОР ДЛЯ ИНЪЕКЦИЙ 0.5%":
        "СИБАЗОН",
    "СИНАФЛАНА МАЗЬ 0.025%":
        "СИНАФЛАН",
    "СТАБИЗОЛ ГЭК 6%":
        "СТАБИЗОЛ ГЭК",
    "СТРЕПТОЦИДОВАЯ МАЗЬ 10%":
        "СТРЕПТОЦИД",
    "СУЛЬФОКАМФОКАИН ДЛЯ ИНЪЕКЦИЙ 10%":
        "СУЛЬФОКАМФОКАИН",
    "ТИАПРИДА ТАБЛЕТКИ 0.1 Г":
        "ТИАПРИД",
    "ФТАЛАЗОЛА ТАБЛЕТКИ 0.5 Г":
        "ФТАЛАЗОЛ",
    "ФУРАГИНА ТАБЛЕТКИ 0.05 Г":
        "ФУРАГИН",
    "ФУРАЗОЛИДОНА ТАБЛЕТКИ 0.05 Г":
        "ФУРАЗОЛИДОН",
    "ФУРАЦИЛИНА ТАБЛЕТКИ ДЛЯ НАРУЖНОГО УПОТРЕБЛЕНИЯ 0.02 Г":
        "ФУРАЦИЛИН",
}


def normalize_brand(raw: str) -> str:
    raw = raw.strip()
    if raw in EXPLICIT_BRAND_MAP:
        return EXPLICIT_BRAND_MAP[raw]
    return raw


# ─────────────────────────────────────────────
# 2.  ProducerTypeScore classification
# ─────────────────────────────────────────────

# Keywords in company name that indicate global pharma group (score 5)
GLOBAL_KEYWORDS = [
    "BAYER", "NOVARTIS", "PFIZER", "ASTRAZENECA", "АСТРАЗЕНЕКА",
    "GLAXOSMITHKLINE", "GLAXO", "SANOFI", "ABBOTT LABORATORIES",
    "JOHNSON & JOHNSON", "MERCK SHARP", "ELI LILLY", "BRISTOL-MYERS",
    "BOEHRINGER INGELHEIM", "FRESENIUS", "B.BRAUN", "BRAUN MELSUNGEN",
    "PROCTER & GAMBLE", "TEVA PHARMACEUTICAL", "SANDOZ INTERNATIONAL",
    "ACTAVIS GROUP", "ALLERGAN", "RECKITT BENCKISER",
    "SWEDISH ORPHAN BIOVITRUM", "SHIRE PLC", "STADA ARZNEIMITTEL",
    "BAUSCH HEALTH",
]

# Keywords for international generics / mid-size pharma (score 4)
INTL_KEYWORDS = [
    "KRKA DD", "LEK PHARMACEUTICALS", "LEK DD", "ЛЕК ДД",
    "GEDEON RICHTER", "EGIS PHARMACEUTICALS",
    "DR.REDDY'S", "DR. REDDY",
    "SUN PHARMACEUTICAL", "GLENMARK", "IPCA LABORATORIES",
    "LUPIN LTD", "PANACEA BIOTEC", "HETERO LABS", "HETERO DRUGS",
    "MEDLEY PHARMACEUTICALS", "CADILA PHARMACEUTICALS",
    "MICRO LABS", "EMCURE", "AUROBINDO PHARMA",
    "MACLEODS", "AJANTA PHARMA", "SIMPLEX PHARMA",
    "UNIQUE PHARMACEUTICAL", "JODAS EXPOIM", "AGIO PHARMACEUTICALS",
    "JIVDHARA PHARMA", "PROTECH BIOSYSTEMS", "OXFORD LABORATORIES",
    "SHREYA LIFE SCIENCES", "VIZAG PHARMACEUTICALS",
    "POLPHARMA SA", "ADAMED PHARMA", "MEDANA PHARMA", "MEDANA ФАРМА",
    "ПОЛЬФАРМА", "POLFA WARSZAWA", "JELFA PHARMACEUTICAL",
    "ALKALOID AD", "BOSNALIJEK", "BELUPO", "PHARMAS DOO",
    "SOPHARMA AD", "DANHSON", "VETPROM",
    "PHARMACEUTICAL BALKANS", "HEMOFARM AD",
    "GRINDEKS AS", "OLAINFARM AS", "SANITAS JSC",
    "BIOCODEX", "A.MENARINI", "DOMPE FARMACEUTICI",
    "ITALFARMACO", "ALFASIGMA", "ANGELINI HOLDING", "S.I.F.I.",
    "PRO.MED.CS", "ZENTIVA GROUP",
    "CHEPLAPHARM", "WORLD MEDICINE LTD", "ROTAPHARM LTD",
    "GALDERMA", "OM PHARMA", "TRB CHEMEDICA",
    "IBSA INSTITUT", "HELSINN HEALTHCARE", "RIVOPHARM",
    "BESINS HEALTHCARE", "PIERRE FABRE",
    "LABORATOIRES MAYOLY", "INNOTECH INTERNATIONAL",
    "TAD PHARMA", "WOERWAG PHARMA", "ARISTO PHARMA",
    "KREWEL MEUSELBACH", "DR.WILLMAR SCHWABE", "DR. FALK",
    "ESPARMA", "MEDAC GMBH", "RATIOPHARM", "EVER PHARMA",
    "LANNACHER", "GEROT PHARMAZEUTIKA",
    "MEGA LIFESCIENCE", "HERBION PAKISTAN",
    "ORION CORPORATION", "MEDA AB",
    "SANOVEL ILAC", "NOBEL ILAC", "WORLD MEDICINE ILAC",
    "BIO-GEN ILAC", "GEN ILAC", "ECZACIBASI",
    "SENJU PHARMACEUTICAL", "SANTEN PHARMACEUTICAL", "SAMIL CO",
    "EMS SA", "ORCHIDIA PHARMACEUTICAL",
    "GENEPHARM SA",
    "VALЕАНТ", "ВАЛЕАНТ", "BAUSCH HEALTH",
    "BALKAN PHARMACEUTICALS",
]

# Top-20 Russian companies by volume → score 3
LARGE_RUSSIAN_KEYWORDS = [
    "ОБНОВЛЕНИЕ ПФК", "ИНФАМЕД", "ОТИСИФАРМ", "ГРОТЕКС", "SOLOPHARM",
    "МОСКОВСКИЙ ЭНДОКРИННЫЙ ЗАВОД", "МЭЗ ФГУП", "ЛЕККО",
    "ЮЖФАРМ", "АТОЛЛ", "ФАРМСТАНДАРТ", "СЕВЕРНАЯ ЗВЕЗДА",
    "УСОЛЬЕ-СИБИРСКИЙ ХИМИКО-ФАРМАЦЕВТИЧЕСКИЙ ЗАВОД",
    "УСОЛЬЕ-СИБИРСКИЙ ХФЗ",
    "ЭВАЛАР", "СИНТЕЗ АКО", "СИНТЕЗ АКО ОАО",
    "АВВА РУС", "ВЕРТЕКС", "ТАТХИМФАРМПРЕПАРАТЫ",
    "ТУЛЬСКАЯ ФАРМАЦЕВТИЧЕСКАЯ ФАБРИКА",
    "НИЖФАРМ", "ОБОЛЕНСКОЕ ФП", "ВАЛЕНТА ФАРМ",
    "ВЕЛФАРМ", "АЛИУМ", "ОЗОН ФАРМАЦЕВТИКА",
    "ФАРМПРОЕКТ", "ВЕРОФАРМ", "АКРИХИН",
    "КАНОНФАРМА", "ФАРМСТАНДАРТ-УФАВИТА",
    "АВЕКСИМА", "ФАРМАСИНТЕЗ", "ПАТЕНТ-ФАРМ",
    "МЕДИСОРБ", "СОТЕКС", "ПОЛИСАН",
    "ГЕРОФАРМ", "ПЕТРОВАКС ФАРМ", "ПРОМОМЕД",
    "БИННОФАРМ", "НАЦИОНАЛЬНАЯ ИССЛЕДОВАТЕЛЬСКАЯ КОМПАНИЯ",
    "ДИАФАРМ", "АЛВИЛС", "АВЕКСИМА",
    "РАНКОФ", "ФЛОРА КАВКАЗА", "ФАРМПРЕПАРАТ",
    "РИФ ООО", "ФАРМВИЛАР", "ДИАСИНТЕЗ",
    "ЛЕКАРЬ ООО", "МЕДСИНТЕЗ", "ПРАНАФАРМ",
    "МАРБИОФАРМ", "РАФАРМА", "ЦИТОМЕД",
    "УРАЛБИОФАРМ", "ФАРМАПЕКС", "Б-ФАРМ",
    "БИОСИНТЕЗ", "БИОХИМИК",
]

# Known city/regional pharma factories → score 1 (local)
FACTORY_KEYWORDS = [
    "ФАРМАЦЕВТИЧЕСКАЯ ФАБРИКА", "ФАРМФАБРИКА", "ФАРМ. ФАБРИКА",
    "ЯРОСЛАВСКАЯ ФАРМАЦЕВТИЧЕСКАЯ", "ИВАНОВСКАЯ ФАРМАЦЕВТИЧЕСКАЯ",
    "КИРОВСКАЯ ФАРМАЦЕВТИЧЕСКАЯ", "САМАРСКАЯ ФАРМАЦЕВТИЧЕСКАЯ",
    "РОСТОВСКАЯ ФАРМАЦЕВТИЧЕСКАЯ", "ВОЛГОГРАДСКАЯ ФАРМАЦЕВТИЧЕСКАЯ",
    "ТВЕРСКАЯ ФАРМАЦЕВТИЧЕСКАЯ", "КЕМЕРОВСКАЯ ФАРМАЦЕВТИЧЕСКАЯ",
    "ОРНОФАРМ", "КРАСФАРМА", "МАBCO", "ВНУТРИАПТЕЧНАЯ ЗАГОТОВКА",
    "БИОВИТ ОСОО",
]

WESTERN_COUNTRIES = {
    "ГЕРМАНИЯ", "ФРАНЦИЯ", "ВЕЛИКОБРИТАНИЯ", "ШВЕЙЦАРИЯ",
    "СОЕДИНЕННЫЕ ШТАТЫ АМЕРИКИ", "КАНАДА", "АВСТРИЯ", "ШВЕЦИЯ",
    "ФИНЛЯНДИЯ", "ИРЛАНДИЯ", "НИДЕРЛАНДЫ", "ИТАЛИЯ", "ИСПАНИЯ",
    "БЕЛЬГИЯ", "ПОРТУГАЛИЯ", "ДАНИЯ", "НОРВЕГИЯ", "МОНАКО",
}

EAST_EU_COUNTRIES = {
    "ВЕНГРИЯ", "РУМЫНИЯ", "ПОЛЬША", "ЧЕХИЯ", "ЧЕШСКАЯ РЕСПУБЛИКА",
    "СЛОВЕНИЯ", "БОЛГАРИЯ", "ХОРВАТИЯ", "СЕРБИЯ", "ЛАТВИЯ",
    "ЛИТВА", "ЭСТОНИЯ", "РЕСПУБЛИКА МАКЕДОНИЯ", "БОСНИЯ И ГЕРЦЕГОВИНА",
    "КИПР", "ЮГОСЛАВИЯ",
}

CIS_COUNTRIES = {
    "БЕЛАРУСЬ", "КАЗАХСТАН", "УКРАИНА", "АРМЕНИЯ",
    "УЗБЕКИСТАН", "МОЛДОВА", "КИРГИЗИЯ",
}


def producer_type_score(firm: str, country: str) -> int:
    fu = firm.upper() if firm else ""
    cu = country.upper().strip() if country else ""

    # Check global pharma keywords
    for kw in GLOBAL_KEYWORDS:
        if kw in fu:
            return 5

    # Special: Israel → Teva is global (5), others are international (4)
    if cu == "ИЗРАИЛЬ":
        if "TEVA" in fu:
            return 5
        return 4

    # Check international generics / mid-size pharma
    for kw in INTL_KEYWORDS:
        if kw in fu:
            return 4

    # Check Japan, South Korea, China – treat as international (3-4)
    if cu in {"ЯПОНИЯ", "ЮЖНАЯ КОРЕЯ"}:
        return 4
    if cu in {"КИТАЙ", "ТАИЛАНД", "БРАЗИЛИЯ", "ЕГИПЕТ", "ПАКИСТАН"}:
        return 3

    # Western countries not yet matched → international (4) by default
    if cu in WESTERN_COUNTRIES:
        return 4

    # Eastern EU not yet matched → regional-international (3)
    if cu in EAST_EU_COUNTRIES:
        return 3

    # India not yet matched → large generic (3)
    if cu == "ИНДИЯ":
        return 3

    # Russia: distinguish by size
    if cu == "РОССИЯ":
        for kw in FACTORY_KEYWORDS:
            if kw in fu:
                return 1
        for kw in LARGE_RUSSIAN_KEYWORDS:
            if kw in fu:
                return 3
        if "НЕИЗВЕСТНЫЙ" in fu or "~" in fu:
            return 0
        return 2  # medium national

    # CIS countries
    if cu in CIS_COUNTRIES:
        # Major CIS manufacturers
        if any(k in fu for k in ["БЕЛМЕДПРЕПАРАТЫ", "БОРИСОВСКИЙ ЗАВОД", "ЛЕКФАРМ", "ФАРМТЕХНОЛОГИЯ",
                                  "БОРЩАГОВСКИЙ", "ФАРМАК", "ДАРНИЦА", "КОНЦЕРН СТИРОЛ",
                                  "ХИМФАРМ", "МИНСКИНТЕРКАПС", "ФЛЮМЕД", "FLUMED"]):
            return 2
        return 1

    # Unknown / ~
    return 0


def coverage_score_from_producer(pt_score: int) -> int:
    """
    CoverageScore: estimated share of world pharma markets where brand is present.
    Derived from producer type as the best available proxy.
    0-5% → 0 (domestic only)
    5-20% → 1
    20-40% → 2
    40-60% → 3
    60-80% → 4
    80-100% → 5
    """
    mapping = {5: 5, 4: 4, 3: 2, 2: 1, 1: 0, 0: 0}
    return mapping.get(pt_score, 0)


# ─────────────────────────────────────────────
# 3.  Scale helpers
# ─────────────────────────────────────────────

def years_score(years_on_market: float) -> int:
    if years_on_market < 1:
        return 0
    elif years_on_market <= 2:
        return 1
    elif years_on_market <= 5:
        return 2
    elif years_on_market <= 10:
        return 3
    elif years_on_market <= 15:
        return 4
    else:
        return 5


def portfolio_score(n: int) -> int:
    if n == 1:
        return 0
    elif n <= 3:
        return 1
    elif n <= 6:
        return 2
    elif n <= 10:
        return 3
    elif n <= 20:
        return 4
    else:
        return 5


# ─────────────────────────────────────────────
# 4.  Main build
# ─────────────────────────────────────────────

print("Loading workbook …")
wb = openpyxl.load_workbook(FILE)
ws = wb["основа_норм"]

# Column indices (1-based for openpyxl, 0-based for tuple access)
C_BRAND = 4    # Торговое наименование  (col 4 = index 3)
C_FIRM  = 8    # Фирма-производитель   (col 8 = index 7)
C_CTRY  = 9    # Страна-производитель  (col 9 = index 8)
C_INN   = 10   # МНН                   (col 10 = index 9)
C_YEAR  = 18   # Год                   (col 18 = index 17)
C_PACKS = 20   # Объем (розница) упак. (col 20 = index 19)
NEW_BRAND_NORM_COL = 29  # Бренд_норм
NEW_BRAND_INDEX_LAG_COL = 30  # BrandIndex_lag1

# Read all data rows as tuples
all_rows = list(ws.iter_rows(min_row=2, values_only=True))
n_rows   = len(all_rows)
print(f"  {n_rows} data rows")

# Helper: safe string
def s(v):
    return str(v).strip() if v else ""

# ── 4a. Normalise brand names ──
brand_norm = [normalize_brand(s(r[C_BRAND - 1])) if r[C_BRAND - 1] else "" for r in all_rows]

# ── 4b. Collect firm metadata (firm → country) ──
firm_country = {}
for r in all_rows:
    firm = s(r[C_FIRM - 1])
    ctry = s(r[C_CTRY - 1])
    if firm and firm not in firm_country:
        firm_country[firm] = ctry

# ── 4c. First year each canonical brand appears (YearsScore proxy) ──
brand_first_year = {}
for i, r in enumerate(all_rows):
    bn = brand_norm[i]
    yr = r[C_YEAR - 1]
    if bn and yr:
        if bn not in brand_first_year or yr < brand_first_year[bn]:
            brand_first_year[bn] = yr

# ── 4d. Total packs per (brand, year) for CoverageScore ──
brand_year_packs = defaultdict(float)
year_total_packs = defaultdict(float)
for i, r in enumerate(all_rows):
    bn = brand_norm[i]
    yr = r[C_YEAR - 1]
    pk = r[C_PACKS - 1] or 0
    if bn and yr:
        brand_year_packs[(bn, yr)] += pk
        year_total_packs[yr] += pk


def coverage_score_from_share(brand: str, year: int) -> int:
    total = year_total_packs.get(year, 0)
    if total == 0:
        return 0
    share_pct = brand_year_packs.get((brand, year), 0) / total * 100
    # Rescaled thresholds derived from actual distribution in dataset
    if share_pct < 0.05:
        return 0
    elif share_pct < 0.2:
        return 1
    elif share_pct < 0.4:
        return 2
    elif share_pct < 0.6:
        return 3
    elif share_pct < 0.8:
        return 4
    else:
        return 5


# ── 4e. Portfolio: unique canonical brand names per (firm, year) ──
firm_year_brands = defaultdict(set)
for i, r in enumerate(all_rows):
    firm = s(r[C_FIRM - 1])
    yr   = r[C_YEAR - 1]
    bn   = brand_norm[i]
    if firm and yr and bn:
        firm_year_brands[(firm, yr)].add(bn)

# ── 4f. Producer type per firm ──
firm_pt = {}
for firm, ctry in firm_country.items():
    firm_pt[firm] = producer_type_score(firm, ctry)

# ─────────────────────────────────────────────
# 5.  Build BrandIndex per (canonical_brand, firm, year)
# ─────────────────────────────────────────────

print("Computing BrandIndex …")

# Collect unique (brand_norm, firm, year) triples
triples_seen = set()
for i, r in enumerate(all_rows):
    bn   = brand_norm[i]
    firm = s(r[C_FIRM - 1])
    yr   = r[C_YEAR - 1]
    if bn and firm and yr:
        triples_seen.add((bn, firm, yr))

# For each triple compute the four components
brand_index_table = []   # list of dicts
for (bn, firm, yr) in sorted(triples_seen):
    first_yr = brand_first_year.get(bn, yr)
    yrs_obs  = yr - first_yr          # lower-bound years on market

    ys  = years_score(yrs_obs)
    cs  = coverage_score_from_share(bn, yr)
    pt  = firm_pt.get(firm, 0)
    cvs = coverage_score_from_producer(pt)   # second coverage variant
    # Use average of both coverage signals (market-share + producer-type)
    cs_final = round((cs + cvs) / 2)
    ps  = portfolio_score(len(firm_year_brands.get((firm, yr), set())))
    bi  = round((ys + cs_final + ps + pt) / 4, 4)

    brand_index_table.append({
        "brand": bn,
        "firm":  firm,
        "year":  yr,
        "YearsScore":       ys,
        "CoverageScore":    cs_final,
        "PortfolioScore":   ps,
        "ProducerTypeScore":pt,
        "BrandIndex":       bi,
    })

# ─────────────────────────────────────────────
# 6.  Write BrandIndex sheet
# ─────────────────────────────────────────────

print("Writing BrandIndex sheet …")
if "BrandIndex" in wb.sheetnames:
    del wb["BrandIndex"]
wbi = wb.create_sheet("BrandIndex")

header = [
    "Торговое наименование (норм.)",
    "Фирма-производитель",
    "Год",
    "YearsScore",
    "CoverageScore",
    "PortfolioScore",
    "ProducerTypeScore",
    "BrandIndex",
]
wbi.append(header)

hdr_font   = Font(bold=True, color="FFFFFF")
hdr_fill   = PatternFill("solid", fgColor="2E4057")
for col_idx, _ in enumerate(header, start=1):
    cell = wbi.cell(1, col_idx)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="center")

for row in brand_index_table:
    wbi.append([
        row["brand"], row["firm"], row["year"],
        row["YearsScore"], row["CoverageScore"],
        row["PortfolioScore"], row["ProducerTypeScore"],
        row["BrandIndex"],
    ])

# Column widths
widths = [45, 55, 8, 12, 15, 14, 17, 12]
for i, w in enumerate(widths, start=1):
    wbi.column_dimensions[get_column_letter(i)].width = w

# ─────────────────────────────────────────────
# 7.  Look-up: BrandIndex for (brand_norm, firm, year) → lag-1
# ─────────────────────────────────────────────

bi_lookup = {}
for row in brand_index_table:
    bi_lookup[(row["brand"], row["firm"], row["year"])] = row["BrandIndex"]

# ─────────────────────────────────────────────
# 8.  Write two new columns to основа_норм
#     Col 29: Бренд_норм
#     Col 30: BrandIndex_lag1
# ─────────────────────────────────────────────

print("Writing new columns to основа_норм …")

# Set header row for new columns
ws.cell(1, NEW_BRAND_NORM_COL).value      = "Бренд_норм"
ws.cell(1, NEW_BRAND_INDEX_LAG_COL).value = "BrandIndex_lag1"

# Style headers to match existing pattern
for col in (NEW_BRAND_NORM_COL, NEW_BRAND_INDEX_LAG_COL):
    c = ws.cell(1, col)
    c.font = Font(bold=True)

for i, r in enumerate(all_rows, start=2):
    bn   = brand_norm[i - 2]
    firm = s(r[C_FIRM - 1])
    yr   = r[C_YEAR - 1]

    ws.cell(i, NEW_BRAND_NORM_COL).value = bn if bn else None

    if bn and firm and yr:
        lag_val = bi_lookup.get((bn, firm, yr - 1))
        ws.cell(i, NEW_BRAND_INDEX_LAG_COL).value = lag_val  # None for first year
    else:
        ws.cell(i, NEW_BRAND_INDEX_LAG_COL).value = None

# ─────────────────────────────────────────────
# 9.  Save
# ─────────────────────────────────────────────

print("Saving …")
wb.save(FILE)
print("Done.")

# Quick stats
print(f"\n=== BrandIndex summary ===")
print(f"  Unique triples (brand_norm × firm × year): {len(brand_index_table)}")
vals = [r["BrandIndex"] for r in brand_index_table]
print(f"  BrandIndex range: {min(vals):.4f} – {max(vals):.4f}")
print(f"  Mean: {sum(vals)/len(vals):.4f}")

coverage_null = sum(1 for i, r in enumerate(all_rows) if
                    r[C_YEAR - 1] and r[C_YEAR - 1] > min(
                        brand_first_year.values(), default=2019)
                    and bi_lookup.get((brand_norm[i], s(r[C_FIRM - 1]), r[C_YEAR - 1] - 1)) is None)
print(f"\n  Rows with BrandIndex_lag1 = None (first year / new brand): {coverage_null}")
non_null = sum(1 for i, r in enumerate(all_rows)
               if ws.cell(i + 2, NEW_BRAND_INDEX_LAG_COL).value is not None)
print(f"  Rows with BrandIndex_lag1 filled: {non_null}")
